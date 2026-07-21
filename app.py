"""
FeeFlow — Complete Fee Management System
========================================
All 26 features implemented:
1.  Batch creation → auto 3 years + 6 sems
2.  Semester fee rules (odd=full, even=hostel/bus only)
3.  Roll format lock after graduation
4.  Student type history tracking
5.  Full student profile popup API
6.  Smart payment form (year/type filtered)
7.  Fee defaulter list + email alerts
8.  Receipt PDF download
9.  Audit log
10. Super-admin manages other admins
11. Batch-wise analytics dashboard
12. Email reminders for due dates
13. Fee waiver/concession
14. No-dues certificate PDF
15. Due date per semester + late fee
16. Payment mode tracking (Cash/DD/UPI/Online/Cheque)
17. Installment plans
18. Scholarship tracking (APCFSS/TSCFSS)
19. Exam eligibility check
20. Transfer Certificate (TC)
21. Student portal (view-only login)
22. College branding on PDFs
23. Excel reports
24. WhatsApp notifications
25. Promotion + undo
26. Admin roles (super_admin/accountant/viewer)
"""

from flask import (Flask, render_template, request, redirect,
                   session, flash, make_response, jsonify,
                   send_file, abort)
from datetime import timedelta, datetime, date
from werkzeug.security import generate_password_hash, check_password_hash
import pymysql
import secrets
import smtplib
import time
import os
import re
import logging
import threading
import io
import json
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from functools import wraps
from collections import defaultdict

# Optional imports — install via requirements.txt
try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    import requests as req_lib
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False

# ─────────────────────────────────────────────
#  APP SETUP
# ─────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', '9ba1b270e638a4f534cc5f66189d2738ee23e79da93aadff8cd458a2e5753e3c')
app.permanent_session_lifetime = timedelta(days=7)

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────
SENDER_EMAIL       = os.environ.get('SENDER_EMAIL',    'kondruamaresh30@gmail.com')
SENDER_PASSWORD    = os.environ.get('SENDER_PASSWORD', 'uams kslw nxkb seng')
WHATSAPP_API_KEY   = os.environ.get('WHATSAPP_API_KEY', '')
WHATSAPP_API_URL   = os.environ.get('WHATSAPP_API_URL', 'https://api.whatsapp-business.example.com/send')

DB_HOST     = os.environ.get('DB_HOST',     'localhost')
DB_USER     = os.environ.get('DB_USER',     'amaresh')
DB_PASSWORD = os.environ.get('DB_PASSWORD', '1234')
DB_NAME     = os.environ.get('DB_NAME',     'fee_management')

OTP_EXPIRY_SECONDS = 300
OTP_MAX_ATTEMPTS   = 3
MIN_PASSWORD_LEN   = 8

# Diploma: year_level -> semester numbers
# Year 1 = Sem 1 ONLY (1 semester, all fees apply)
# Year 2 = Sem 3 & 4  (Sem 3 full fees; Sem 4 hostel/bus only)
# Year 3 = Sem 5 & 6  (Sem 5 full fees; Sem 6 hostel/bus only)
DIPLOMA_SEMESTERS = {1: [1], 2: [3, 4], 3: [5, 6]}

# Semesters that get full fees (college + building + hostel/bus)
# Sem 1 is Year-1's only semester → ALL fees
FULL_FEE_SEMS  = {1, 3, 5}
# Semesters that get ONLY hostel/bus fees (even sems of Year 2 & 3)
HOSTEL_BUS_ONLY_SEMS = {4, 6}

# ─────────────────────────────────────────────
#  RATE LIMITER
# ─────────────────────────────────────────────
_rate_data = defaultdict(list)
_rate_lock = threading.Lock()

def _is_rate_limited(key, max_calls, window):
    now = time.time()
    with _rate_lock:
        _rate_data[key] = [t for t in _rate_data[key] if now - t < window]
        if len(_rate_data[key]) >= max_calls:
            return True
        _rate_data[key].append(now)
        return False

def rate_limit(max_calls, window):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            key = f"{request.remote_addr}:{f.__name__}"
            if _is_rate_limited(key, max_calls, window):
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify(error="Too many requests. Please wait."), 429
                flash("Too many requests. Please wait.", "error")
                return redirect(request.referrer or '/')
            return f(*args, **kwargs)
        return wrapped
    return decorator

# ─────────────────────────────────────────────
#  DATABASE
# ─────────────────────────────────────────────
def get_db():
    return pymysql.connect(
        host=DB_HOST, user=DB_USER, password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor,
        connect_timeout=5,
        autocommit=False,
    )

# ─────────────────────────────────────────────
#  CSRF
# ─────────────────────────────────────────────
def generate_csrf_token():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']

app.jinja_env.globals['csrf_token'] = generate_csrf_token

# ─────────────────────────────────────────────
#  AUDIT LOG
# ─────────────────────────────────────────────
def audit(action, target_type=None, target_id=None, details=None):
    try:
        conn = get_db()
        with conn.cursor() as c:
            c.execute("""
                INSERT INTO audit_log
                    (admin_id, admin_name, action, target_type, target_id, details, ip_address)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                session.get('user_id'),
                session.get('user', 'system'),
                action,
                target_type,
                target_id,
                json.dumps(details) if details and not isinstance(details, str) else details,
                request.remote_addr if request else None
            ))
            conn.commit()
    except Exception as e:
        logger.error("Audit log error: %s", e)
    finally:
        try: conn.close()
        except: pass

# ─────────────────────────────────────────────
#  OTP HELPERS
# ─────────────────────────────────────────────
def _otp_save(email, otp):
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                INSERT INTO otp_store (email, otp, created_at, attempts)
                VALUES (%s, %s, NOW(), 0)
                ON DUPLICATE KEY UPDATE otp=%s, created_at=NOW(), attempts=0
            """, (email, otp, otp))
        conn.commit()
    finally: conn.close()

def _otp_get(email):
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT * FROM otp_store WHERE email=%s", (email,))
            return c.fetchone()
    finally: conn.close()

def _otp_increment(email):
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("UPDATE otp_store SET attempts=attempts+1 WHERE email=%s", (email,))
        conn.commit()
    finally: conn.close()

def _otp_delete(email):
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("DELETE FROM otp_store WHERE email=%s", (email,))
        conn.commit()
    finally: conn.close()

def _generate_otp(length=6):
    return ''.join(str(secrets.randbelow(10)) for _ in range(length))

# ─────────────────────────────────────────────
#  EMAIL
# ─────────────────────────────────────────────
def _smtp_send(msg):
    with smtplib.SMTP("smtp.gmail.com", 587, timeout=10) as s:
        s.ehlo(); s.starttls()
        s.login(SENDER_EMAIL, SENDER_PASSWORD)
        s.send_message(msg)

def send_otp_email(to_email, otp):
    if not SENDER_EMAIL or not SENDER_PASSWORD:
        logger.warning("Email not configured — OTP: %s", otp)
        return True
    try:
        msg = MIMEText(
            f"Your OTP for password reset is: {otp}\n\n"
            f"This code expires in {OTP_EXPIRY_SECONDS // 60} minutes.\n"
            "If you did not request this, please ignore this email."
        )
        msg['Subject'] = "Password Reset OTP — FeeFlow"
        msg['From'] = SENDER_EMAIL
        msg['To'] = to_email
        _smtp_send(msg)
        return True
    except Exception as e:
        logger.error("OTP email failed: %s", e)
        return False

def send_receipt_email(to_email, name, roll_no, receipt_no, fee_type,
                       amount, pay_date, paid_by, payment_mode='cash',
                       pdf_bytes=None):
    if not SENDER_EMAIL or not SENDER_PASSWORD:
        return True
    try:
        msg = MIMEMultipart()
        msg['Subject'] = f"Fee Payment Receipt — {receipt_no}"
        msg['From'] = SENDER_EMAIL
        msg['To'] = to_email
        body = (
            f"Dear {name},\n\n"
            f"Your fee payment has been recorded successfully.\n\n"
            f"Receipt No   : {receipt_no}\n"
            f"Roll No      : {roll_no}\n"
            f"Fee Type     : {fee_type}\n"
            f"Amount Paid  : Rs.{amount:,.2f}\n"
            f"Payment Mode : {payment_mode.upper()}\n"
            f"Date         : {pay_date}\n"
            f"Recorded By  : {paid_by}\n\n"
            f"Please keep this for your records.\n\n— FeeFlow System"
        )
        msg.attach(MIMEText(body))
        if pdf_bytes:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(pdf_bytes)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="receipt_{receipt_no}.pdf"')
            msg.attach(part)
        _smtp_send(msg)
        return True
    except Exception as e:
        logger.error("Receipt email failed: %s", e)
        return False

def send_reminder_email(to_email, name, roll_no, sem_name, due_date, pending_amount):
    if not SENDER_EMAIL or not SENDER_PASSWORD:
        return True
    try:
        msg = MIMEText(
            f"Dear {name},\n\n"
            f"This is a reminder that your fee payment for {sem_name} is due.\n\n"
            f"Roll No          : {roll_no}\n"
            f"Semester         : {sem_name}\n"
            f"Due Date         : {due_date}\n"
            f"Pending Amount   : Rs.{pending_amount:,.2f}\n\n"
            f"Please pay at the earliest to avoid late fees.\n\n— FeeFlow System"
        )
        msg['Subject'] = f"Fee Reminder — {sem_name} Due {due_date}"
        msg['From'] = SENDER_EMAIL
        msg['To'] = to_email
        _smtp_send(msg)
        return True
    except Exception as e:
        logger.error("Reminder email failed: %s", e)
        return False

# ─────────────────────────────────────────────
#  WHATSAPP
# ─────────────────────────────────────────────
def send_whatsapp(phone, message):
    if not WHATSAPP_API_KEY or not REQUESTS_OK:
        logger.info("WhatsApp not configured. Message: %s", message[:50])
        return False
    try:
        resp = req_lib.post(WHATSAPP_API_URL, json={
            'api_key': WHATSAPP_API_KEY,
            'phone': phone,
            'message': message
        }, timeout=8)
        return resp.status_code == 200
    except Exception as e:
        logger.error("WhatsApp send failed: %s", e)
        return False

# ─────────────────────────────────────────────
#  AUTH DECORATOR
# ─────────────────────────────────────────────
ROLE_HIERARCHY = {
    'super_admin': 5,
    'admin': 4,
    'accountant': 3,
    'viewer': 2,
    'student': 1,
}

def login_required(role=None, min_role=None):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if 'user' not in session:
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify(error="Authentication required."), 401
                flash("Your session has expired. Please log in again.", "warning")
                return redirect('/')
            user_role = session.get('role', '')
            if role and user_role != role:
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify(error="Forbidden."), 403
                flash("Access denied.", "error")
                return redirect('/')
            if min_role:
                if ROLE_HIERARCHY.get(user_role, 0) < ROLE_HIERARCHY.get(min_role, 0):
                    if request.is_json or request.path.startswith('/api/'):
                        return jsonify(error="Insufficient permissions."), 403
                    flash("Access denied.", "error")
                    return redirect('/')
            return f(*args, **kwargs)
        return wrapped
    return decorator

def admin_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if 'user' not in session:
            if request.is_json or request.path.startswith('/api/'):
                return jsonify(error="Authentication required."), 401
            return redirect('/')
        role = session.get('role', '')
        if ROLE_HIERARCHY.get(role, 0) < ROLE_HIERARCHY.get('viewer', 0):
            if request.is_json or request.path.startswith('/api/'):
                return jsonify(error="Forbidden."), 403
            return redirect('/')
        return f(*args, **kwargs)
    return wrapped

def write_required(f):
    """Requires at least accountant role."""
    @wraps(f)
    def wrapped(*args, **kwargs):
        if 'user' not in session:
            return jsonify(error="Authentication required."), 401
        role = session.get('role', '')
        if ROLE_HIERARCHY.get(role, 0) < ROLE_HIERARCHY.get('accountant', 0):
            return jsonify(error="Read-only access. Contact super admin."), 403
        return f(*args, **kwargs)
    return wrapped

def super_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if session.get('role') not in ('super_admin', 'admin'):
            return jsonify(error="Super admin access required."), 403
        return f(*args, **kwargs)
    return wrapped

# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
def _require(data, *fields):
    missing = [f for f in fields if not data.get(f)]
    if missing:
        return jsonify(error=f"Missing required fields: {', '.join(missing)}"), 400
    return None

def is_strong_password(pw):
    if len(pw) < MIN_PASSWORD_LEN:
        return False, f"Password must be at least {MIN_PASSWORD_LEN} characters."
    if not re.search(r'[A-Z]', pw): return False, "Must contain an uppercase letter."
    if not re.search(r'[a-z]', pw): return False, "Must contain a lowercase letter."
    if not re.search(r'\d', pw):    return False, "Must contain a digit."
    if not re.search(r'[!@#$%^&*(),.?\":{}|<>_\-]', pw):
        return False, "Must contain a special character."
    return True, ''

def get_college_settings():
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT * FROM college_settings LIMIT 1")
            return c.fetchone() or {}
    finally: conn.close()

# ─────────────────────────────────────────────
#  PDF GENERATION
# ─────────────────────────────────────────────
def generate_receipt_pdf(payment_data, student_data, college):
    if not REPORTLAB_OK:
        return None
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=15*mm, leftMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story = []

    # Header
    header_style = ParagraphStyle('header', fontSize=16, alignment=TA_CENTER,
                                   fontName='Helvetica-Bold', spaceAfter=4)
    sub_style = ParagraphStyle('sub', fontSize=10, alignment=TA_CENTER,
                                fontName='Helvetica', spaceAfter=2)
    receipt_style = ParagraphStyle('receipt', fontSize=13, alignment=TA_CENTER,
                                    fontName='Helvetica-Bold', spaceBefore=8, spaceAfter=8)

    story.append(Paragraph(college.get('college_name', 'College Name'), header_style))
    if college.get('address'):
        story.append(Paragraph(college['address'], sub_style))
    if college.get('phone') or college.get('email'):
        contact = ' | '.join(filter(None, [college.get('phone'), college.get('email')]))
        story.append(Paragraph(contact, sub_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1a3a6b')))
    story.append(Paragraph("FEE PAYMENT RECEIPT", receipt_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#cccccc')))
    story.append(Spacer(1, 6*mm))

    # Student info table
    info_data = [
        ['Receipt No.', payment_data.get('receipt_no', '—'),
         'Date', str(payment_data.get('date', '—'))],
        ['Student Name', student_data.get('name', '—'),
         'Roll No.', student_data.get('roll_no', '—')],
        ['Branch', student_data.get('branch', '—'),
         'Year', f"Year {student_data.get('year_level', '—')}"],
        ['Student Type', (student_data.get('student_type') or '—').replace('_', ' ').title(),
         'Payment Mode', (payment_data.get('payment_mode') or 'cash').upper()],
    ]
    info_table = Table(info_data, colWidths=[40*mm, 65*mm, 35*mm, 45*mm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f4ff')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#f0f4ff')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#fafafa')]),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 8*mm))

    # Fee detail
    fee_data = [['Fee Type', 'Semester', 'Amount (Rs.)']]
    fee_data.append([
        payment_data.get('fee_type', '—'),
        payment_data.get('sem_name', '—'),
        f"Rs. {float(payment_data.get('amount', 0)):,.2f}",
    ])
    fee_data.append(['', 'Total Paid', f"Rs. {float(payment_data.get('amount', 0)):,.2f}"])

    fee_table = Table(fee_data, colWidths=[80*mm, 60*mm, 45*mm])
    fee_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTNAME', (1, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a6b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f4e8')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(fee_table)
    story.append(Spacer(1, 12*mm))

    # Footer
    footer_style = ParagraphStyle('footer', fontSize=8, alignment=TA_CENTER,
                                   textColor=colors.grey)
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#cccccc')))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        f"Recorded by: {payment_data.get('recorded_by', '—')} | "
        f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')} | "
        "This is a computer-generated receipt.",
        footer_style
    ))
    if college.get('principal_name'):
        story.append(Spacer(1, 8*mm))
        sig_data = [['', 'For ' + college.get('college_name', '')],
                    ['Student Signature', college.get('principal_name', '') + '\nPrincipal']]
        sig_table = Table(sig_data, colWidths=[90*mm, 95*mm])
        sig_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ]))
        story.append(sig_table)

    doc.build(story)
    return buf.getvalue()

def generate_no_dues_pdf(student_data, college):
    if not REPORTLAB_OK:
        return None
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=20*mm, leftMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    story = []

    h1 = ParagraphStyle('h1', fontSize=18, alignment=TA_CENTER, fontName='Helvetica-Bold')
    h2 = ParagraphStyle('h2', fontSize=13, alignment=TA_CENTER, fontName='Helvetica-Bold',
                         spaceBefore=8, spaceAfter=8)
    normal_c = ParagraphStyle('nc', fontSize=11, alignment=TA_CENTER, spaceAfter=6)
    normal_l = ParagraphStyle('nl', fontSize=10, alignment=TA_LEFT, spaceAfter=4)

    story.append(Paragraph(college.get('college_name', 'College Name'), h1))
    if college.get('address'):
        story.append(Paragraph(college['address'], normal_c))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1a6b3a')))
    story.append(Spacer(1, 8*mm))
    story.append(Paragraph("NO DUES CERTIFICATE", h2))
    story.append(Spacer(1, 6*mm))

    story.append(Paragraph(
        f"This is to certify that <b>{student_data.get('name', '—')}</b> "
        f"(Roll No: <b>{student_data.get('roll_no', '—')}</b>, "
        f"Branch: <b>{student_data.get('branch', '—')}</b>) "
        f"has cleared <b>ALL</b> dues and has no pending fees as on "
        f"<b>{datetime.now().strftime('%d-%m-%Y')}</b>.",
        normal_l
    ))
    story.append(Spacer(1, 10*mm))

    if college.get('principal_name'):
        story.append(Spacer(1, 12*mm))
        story.append(Paragraph(
            f"<b>{college['principal_name']}</b><br/>Principal",
            ParagraphStyle('sig', fontSize=10, alignment=TA_RIGHT)
        ))

    doc.build(story)
    return buf.getvalue()

# ─────────────────────────────────────────────
#  LOGIN / LOGOUT
# ─────────────────────────────────────────────
@app.route('/', methods=['GET', 'POST'])
@rate_limit(max_calls=20, window=60)
def login():
    if 'user' in session and session.get('role') in ('super_admin', 'admin', 'accountant', 'viewer'):
        return redirect('/admin')
    if 'user' in session and session.get('role') == 'student':
        return redirect('/student')

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        remember = request.form.get('remember')

        if not username or not password:
            flash("Username and password are required.", "error")
            return render_template('login.html')

        conn = get_db()
        try:
            with conn.cursor() as c:
                # Support login by email OR username (UI sends email as identity)
                c.execute(
                    "SELECT id, username, password, role, email FROM users "
                    "WHERE username=%s OR email=%s",
                    (username, username.lower())
                )
                user = c.fetchone()
        finally: conn.close()

        dummy  = generate_password_hash("dummy-value-anti-timing")
        stored = user['password'] if user else dummy

        valid_roles = ('super_admin', 'admin', 'accountant', 'viewer', 'student')
        if user and user['role'] in valid_roles and check_password_hash(stored, password):
            session.clear()
            session['user']       = user['username']
            session['user_id']    = user['id']
            session['role']       = user['role']
            session['email']      = user.get('email', '')
            session['login_time'] = time.time()
            session.permanent     = bool(remember)
            generate_csrf_token()
            audit('LOGIN', 'user', user['id'])
            if user['role'] == 'student':
                return redirect('/student')
            return redirect('/admin')

        logger.warning("Failed login for username='%s'", username)
        flash("Invalid username or password.", "error")
        return render_template('login.html')

    return render_template('login.html')


@app.route('/logout')
def logout():
    audit('LOGOUT', 'user', session.get('user_id'))
    session.clear()
    resp = make_response(redirect('/'))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    flash("Logged out successfully.", "info")
    return resp

# ─────────────────────────────────────────────
#  ADMIN DASHBOARD
# ─────────────────────────────────────────────
@app.route('/admin')
@admin_required
def admin_dashboard():
    return render_template('admin.html',
                           admin_name=session.get('user', 'Admin'),
                           role=session.get('role', 'viewer'))

# ─────────────────────────────────────────────
#  API — DASHBOARD STATS
# ─────────────────────────────────────────────
# REPLACE the entire admin_dashboard_data function with this:

@app.route('/api/admin/dashboard')
@admin_required
def admin_dashboard_data():
    conn = get_db()
    try:
        with conn.cursor() as c:
            # Student counts
            c.execute("SELECT COUNT(*) AS v FROM users WHERE role='student' AND status='active'")
            total_students = c.fetchone()['v']

            c.execute("SELECT COUNT(*) AS v FROM users WHERE role='student' AND student_type='hosteler' AND status='active'")
            hostelers = c.fetchone()['v']

            c.execute("SELECT COUNT(*) AS v FROM users WHERE role='student' AND student_type='day_scholar' AND status='active'")
            day_scholars = c.fetchone()['v']

            # Total assigned fees (net of waivers/scholarships)
            c.execute("""
                SELECT IFNULL(SUM(sf.amount - sf.waiver_amount - sf.scholarship_amount), 0) AS v
                FROM student_fees sf
                JOIN users u ON sf.student_id = u.id
                WHERE u.role = 'student' AND u.status = 'active'
            """)
            total_net_due = float(c.fetchone()['v'])

            # Total collected
            c.execute("""
                SELECT IFNULL(SUM(amount), 0) AS v 
                FROM payments WHERE status='Paid'
            """)
            collected = float(c.fetchone()['v'])

            # Outstanding = total assigned - collected
            outstanding = max(0.0, total_net_due - collected)

            # Today's stats
            c.execute("""
                SELECT IFNULL(SUM(amount), 0) AS v 
                FROM payments WHERE status='Paid' AND DATE(date) = CURDATE()
            """)
            collected_today = float(c.fetchone()['v'])

            c.execute("""
                SELECT COUNT(*) AS v 
                FROM payments WHERE status='Paid' AND DATE(date) = CURDATE()
            """)
            receipts_today = c.fetchone()['v']

            # Defaulters
            c.execute("""
                SELECT COUNT(DISTINCT sf.student_id) AS v
                FROM student_fees sf
                JOIN fee_structure fs ON sf.fee_structure_id = fs.id
                JOIN semesters s ON fs.semester_id = s.id
                JOIN users u ON sf.student_id = u.id
                WHERE u.status = 'active'
                AND s.due_date IS NOT NULL
                AND s.due_date < CURDATE()
                AND (sf.amount - sf.waiver_amount - sf.scholarship_amount) >
                    IFNULL((
                        SELECT SUM(p.amount) FROM payments p
                        WHERE p.student_id = sf.student_id
                        AND p.fee_structure_id = sf.fee_structure_id
                        AND p.status = 'Paid'
                    ), 0)
            """)
            defaulters = c.fetchone()['v']

    finally:
        conn.close()

    return jsonify(
        total_students  = total_students,
        collected       = collected,
        pending         = outstanding,
        collected_today = collected_today,
        receipts_today  = receipts_today,
        hostelers       = hostelers,
        day_scholars    = day_scholars,
        defaulters      = defaulters,
    )

# ─────────────────────────────────────────────
#  API — BATCH ANALYTICS
# ─────────────────────────────────────────────
@app.route('/api/admin/batch_analytics')
@admin_required
def batch_analytics():
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT * FROM batches ORDER BY batch_year DESC")
            batches = c.fetchall()
            result = []
            for b in batches:
                # Fix: query only needs batch_id once
                c.execute("""
                    SELECT
                        COUNT(DISTINCT u.id) AS total,
                        IFNULL(SUM(sf.amount - sf.waiver_amount - sf.scholarship_amount), 0) AS total_due
                    FROM users u
                    LEFT JOIN student_fees sf ON sf.student_id = u.id
                    WHERE u.batch_id = %s AND u.role = 'student'
                """, (b['id'],))
                stats = c.fetchone()

                c.execute("""
                    SELECT IFNULL(SUM(p.amount), 0) AS collected
                    FROM payments p
                    JOIN users u ON p.student_id = u.id
                    WHERE u.batch_id = %s AND p.status = 'Paid'
                """, (b['id'],))
                pay_stats = c.fetchone()

                total_due = float(stats['total_due'] or 0)
                collected = float(pay_stats['collected'] or 0)
                pct = round(collected / total_due * 100, 1) if total_due > 0 else 0

                result.append({
                    'batch_id':     b['id'],
                    'batch_year':   b['batch_year'],
                    'college_code': b['college_code'],
                    'status':       b['status'],
                    'students':     stats['total'],
                    'total_due':    total_due,
                    'collected':    collected,
                    'pending':      max(0, total_due - collected),
                    'pct':          pct,
                })
    finally:
        conn.close()
    return jsonify(result)

# ─────────────────────────────────────────────
#  API — BATCHES
# ─────────────────────────────────────────────
@app.route('/api/batches', methods=['GET'])
@admin_required
def get_batches():
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT * FROM batches ORDER BY batch_year DESC")
            return jsonify(c.fetchall())
    finally: conn.close()


@app.route('/api/batches', methods=['POST'])
@write_required
def create_batch():
    data         = request.json or {}
    batch_year   = int(data.get('batch_year') or 0)
    college_code = str(data.get('college_code') or '').strip()

    if not batch_year or not college_code:
        return jsonify(error="batch_year and college_code are required."), 400
    if batch_year < 2000 or batch_year > 2100:
        return jsonify(error="Invalid batch year."), 400

    batch_code = str(batch_year)[-2:]

    conn = get_db()
    try:
        with conn.cursor() as c:
            # Check graduated lock
            c.execute("""
                SELECT id FROM graduated_batch_formats
                WHERE batch_code=%s AND college_code=%s
            """, (batch_code, college_code))
            if c.fetchone():
                return jsonify(
                    error=f"Batch code '{batch_code}' with college '{college_code}' "
                          f"is permanently locked due to a graduated student."
                ), 409

            # Check duplicate batch
            c.execute("""
                SELECT id FROM batches 
                WHERE batch_year=%s AND college_code=%s
            """, (batch_year, college_code))
            if c.fetchone():
                return jsonify(
                    error=f"Batch {batch_year} for college '{college_code}' already exists."
                ), 409

            # Create batch
            c.execute("""
                INSERT INTO batches (batch_year, batch_code, college_code, status, created_by)
                VALUES (%s, %s, %s, 'active', %s)
            """, (batch_year, batch_code, college_code, session.get('user_id')))
            batch_id = conn.insert_id()

            # Create 3 academic years + semesters
            year_sems = {1: [1], 2: [3, 4], 3: [5, 6]}
            for yr_order in (1, 2, 3):
                y1 = batch_year + yr_order - 1
                y2 = y1 + 1
                year_name = f"{y1}-{str(y2)[-2:]}"
                is_active = 1 if yr_order == 1 else 0

                # Check if this year_name already exists for THIS batch
                # (shouldn't happen but be safe)
                c.execute("""
                    SELECT id FROM academic_years 
                    WHERE year_name=%s AND batch_id=%s
                """, (year_name, batch_id))
                if c.fetchone():
                    continue  # skip if already exists

                c.execute("""
                    INSERT INTO academic_years 
                        (year_name, is_active, batch_id, year_order)
                    VALUES (%s, %s, %s, %s)
                """, (year_name, is_active, batch_id, yr_order))
                ay_id = conn.insert_id()

                for sem_no in year_sems[yr_order]:
                    c.execute("""
                        INSERT INTO semesters (academic_year_id, semester_no)
                        VALUES (%s, %s)
                    """, (ay_id, sem_no))

        conn.commit()
        audit('CREATE_BATCH', 'batch', batch_id,
              {'batch_year': batch_year, 'college_code': college_code})
        return jsonify(
            success=True,
            batch_id=batch_id,
            message=f"Batch {batch_year} created with 3 academic years and 6 semesters."
        )

    except Exception as e:
        conn.rollback()
        logger.error("Create batch error: %s", e)
        return jsonify(error=f"Failed to create batch: {str(e)}"), 500
    finally:
        conn.close()

# ─────────────────────────────────────────────
#  API — STUDENTS
# ─────────────────────────────────────────────
@app.route('/api/admin/students')
@admin_required
def get_students():
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT u.id, u.roll_no, u.name, u.branch, u.year_level,
                       u.status, u.email, u.student_type, u.transport_type,
                       u.parent_phone, u.batch_id, u.tc_issued,
                       b.batch_year, b.college_code
                FROM users u
                LEFT JOIN batches b ON u.batch_id = b.id
                WHERE u.role='student'
                ORDER BY u.roll_no ASC
            """)
            return jsonify(c.fetchall())
    finally: conn.close()


@app.route('/api/admin/student_profile/<int:student_id>')
@admin_required
def student_full_profile(student_id):
    """Full profile for double-click popup — all history, fees, payments, type changes."""
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT u.*, b.batch_year, b.college_code
                FROM users u
                LEFT JOIN batches b ON u.batch_id=b.id
                WHERE u.id=%s AND u.role='student'
            """, (student_id,))
            student = c.fetchone()
            if not student:
                return jsonify(error="Student not found."), 404

            # Fee summary
            c.execute("""
                SELECT
                    sf.fee_structure_id,
                    fs.fee_name,
                    fs.fee_category,
                    s.semester_no,
                    ay.year_name,
                    sf.amount AS total,
                    sf.waiver_amount,
                    sf.scholarship_amount,
                    (sf.amount - sf.waiver_amount - sf.scholarship_amount) AS net_due,
                    IFNULL((SELECT SUM(p2.amount) FROM payments p2
                            WHERE p2.student_id=sf.student_id
                            AND p2.fee_structure_id=sf.fee_structure_id
                            AND p2.status='Paid'),0) AS paid,
                    s.due_date
                FROM student_fees sf
                JOIN fee_structure fs ON sf.fee_structure_id=fs.id
                JOIN semesters s ON fs.semester_id=s.id
                JOIN academic_years ay ON s.academic_year_id=ay.id
                WHERE sf.student_id=%s
                ORDER BY s.semester_no, fs.fee_category
            """, (student_id,))
            fees = c.fetchall()
            for f in fees:
                f['total']   = float(f['total'])
                f['paid']    = float(f['paid'])
                f['net_due'] = float(f['net_due'])
                f['due']     = round(f['net_due'] - f['paid'], 2)
                if f.get('due_date'): f['due_date'] = str(f['due_date'])

            # Payment history
            c.execute("""
                SELECT p.*, fs.fee_name, s.semester_no, ay.year_name
                FROM payments p
                JOIN fee_structure fs ON p.fee_structure_id=fs.id
                JOIN semesters s ON fs.semester_id=s.id
                JOIN academic_years ay ON s.academic_year_id=ay.id
                WHERE p.student_id=%s AND p.status='Paid'
                ORDER BY p.date DESC, p.id DESC
            """, (student_id,))
            payments = c.fetchall()
            for p in payments:
                p['amount'] = float(p['amount'])
                if p.get('date'): p['date'] = str(p['date'])

            # Type history
            c.execute("""
                SELECT * FROM student_type_history
                WHERE student_id=%s ORDER BY from_date DESC
            """, (student_id,))
            type_history = c.fetchall()
            for t in type_history:
                if t.get('from_date'): t['from_date'] = str(t['from_date'])
                if t.get('to_date'):   t['to_date']   = str(t['to_date'])

            # Scholarships
            c.execute("""
                SELECT sch.*, fs.fee_name, s.semester_no
                FROM scholarships sch
                JOIN fee_structure fs ON sch.fee_structure_id=fs.id
                JOIN semesters s ON fs.semester_id=s.id
                WHERE sch.student_id=%s ORDER BY sch.created_at DESC
            """, (student_id,))
            scholarships = c.fetchall()
            for s in scholarships:
                s['amount'] = float(s['amount'])

            # Waivers
            c.execute("""
                SELECT fw.*, fs.fee_name FROM fee_waivers fw
                JOIN fee_structure fs ON fw.fee_structure_id=fs.id
                WHERE fw.student_id=%s
            """, (student_id,))
            waivers = c.fetchall()
            for w in waivers:
                w['waiver_amount'] = float(w['waiver_amount'])

            # TC record
            c.execute("SELECT * FROM tc_records WHERE student_id=%s", (student_id,))
            tc = c.fetchone()
            if tc and tc.get('issued_at'): tc['issued_at'] = str(tc['issued_at'])

    finally: conn.close()

    return jsonify(
        student      = student,
        fees         = fees,
        payments     = payments,
        type_history = type_history,
        scholarships = scholarships,
        waivers      = waivers,
        tc           = tc,
    )


@app.route('/api/admin/student_by_roll')
@admin_required
def student_by_roll():
    roll = request.args.get('roll', '').strip()
    if not roll:
        return jsonify(error='Roll number is required.'), 400
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT u.id, u.roll_no, u.name, u.branch, u.year_level,
                       u.status, u.email, u.student_type, u.transport_type,
                       u.batch_id, b.batch_year
                FROM users u
                LEFT JOIN batches b ON u.batch_id=b.id
                WHERE u.roll_no=%s AND u.role='student'
            """, (roll,))
            student = c.fetchone()
            if not student:
                return jsonify(error=f'No student found with roll number "{roll}"'), 404

            # Get fees for this student filtered by current year
            sem_nos = DIPLOMA_SEMESTERS.get(student['year_level'], [])
            if sem_nos:
                ph = ','.join(['%s'] * len(sem_nos))
                c.execute(f"""
                    SELECT
                        sf.fee_structure_id,
                        fs.fee_name,
                        fs.fee_category,
                        s.semester_no,
                        ay.year_name,
                        sf.amount AS total,
                        (sf.amount - sf.waiver_amount - sf.scholarship_amount) AS net_due,
                        IFNULL(SUM(CASE WHEN p.status='Paid' THEN p.amount ELSE 0 END),0) AS paid,
                        (sf.amount - sf.waiver_amount - sf.scholarship_amount -
                         IFNULL(SUM(CASE WHEN p.status='Paid' THEN p.amount ELSE 0 END),0)) AS due
                    FROM student_fees sf
                    JOIN fee_structure fs ON sf.fee_structure_id=fs.id
                    JOIN semesters s ON fs.semester_id=s.id
                    JOIN academic_years ay ON s.academic_year_id=ay.id
                    LEFT JOIN payments p ON p.student_id=sf.student_id
                        AND p.fee_structure_id=sf.fee_structure_id
                    WHERE sf.student_id=%s AND s.semester_no IN ({ph})
                    GROUP BY sf.fee_structure_id, fs.fee_name, fs.fee_category,
                             s.semester_no, ay.year_name, sf.amount,
                             sf.waiver_amount, sf.scholarship_amount
                    ORDER BY s.semester_no
                """, (student['id'], *sem_nos))
                fees = c.fetchall()
                for f in fees:
                    f['total']   = float(f['total'])
                    f['paid']    = float(f['paid'])
                    f['net_due'] = float(f['net_due'])
                    f['due']     = round(float(f['due']), 2)
            else:
                fees = []
    finally: conn.close()
    return jsonify(student=student, fees=fees)


@app.route('/api/admin/add_student', methods=['POST'])
@write_required
def add_student():
    data           = request.json or {}
    err            = _require(data, 'roll_no', 'name', 'branch', 'year_level', 'student_type')
    if err: return err

    roll_no        = str(data['roll_no']).strip()
    name           = str(data['name']).strip()
    branch         = str(data['branch']).strip()
    year_level     = int(data['year_level'])
    student_type   = str(data['student_type']).strip()
    transport_type = str(data.get('transport_type') or '').strip() or None
    email          = str(data.get('email') or '').strip() or None
    parent_phone   = str(data.get('parent_phone') or '').strip() or None
    batch_id       = data.get('batch_id')

    if year_level not in (1, 2, 3):
        return jsonify(error="year_level must be 1, 2, or 3."), 400
    if student_type not in ('hosteler', 'day_scholar'):
        return jsonify(error="student_type must be 'hosteler' or 'day_scholar'."), 400
    if student_type == 'day_scholar' and transport_type not in ('own_transport', 'college_bus'):
        return jsonify(error="Day scholar requires transport_type."), 400

    # Roll format lock check
    lock_err = _check_roll_lock(roll_no)
    if lock_err: return lock_err

    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT id FROM users WHERE roll_no=%s OR username=%s", (roll_no, roll_no))
            if c.fetchone():
                return jsonify(error="Roll number already exists."), 409

            c.execute("""
                INSERT INTO users
                    (roll_no, name, branch, username, email, year_level,
                     password, role, status, student_type, transport_type,
                     batch_id, parent_phone, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,'student','active',%s,%s,%s,%s,%s)
            """, (
                roll_no, name, branch, roll_no, email, year_level,
                generate_password_hash(roll_no),
                student_type,
                transport_type if student_type == 'day_scholar' else None,
                batch_id, parent_phone, session.get('user_id')
            ))
            user_id = conn.insert_id()

            # Record initial type history
            c.execute("""
                INSERT INTO student_type_history
                    (student_id, student_type, transport_type, from_date, changed_by)
                VALUES (%s,%s,%s,CURDATE(),%s)
            """, (user_id, student_type,
                  transport_type if student_type == 'day_scholar' else None,
                  session.get('user')))
        conn.commit()
        _auto_assign_fees([{'id': user_id, 'roll_no': roll_no}],
                          year_level, student_type, transport_type, batch_id)
        audit('ADD_STUDENT', 'student', user_id, {'roll_no': roll_no, 'name': name})
    finally: conn.close()

    return jsonify(success=True, user_id=user_id)


@app.route('/api/admin/edit_student/<int:student_id>', methods=['PUT'])
@write_required
def edit_student(student_id):
    data           = request.json or {}
    err            = _require(data, 'name', 'branch', 'year_level', 'student_type')
    if err: return err

    name           = str(data['name']).strip()
    branch         = str(data['branch']).strip()
    year_level     = int(data['year_level'])
    student_type   = str(data['student_type']).strip()
    transport_type = str(data.get('transport_type') or '').strip() or None
    email          = str(data.get('email') or '').strip() or None
    parent_phone   = str(data.get('parent_phone') or '').strip() or None

    if student_type not in ('hosteler', 'day_scholar'):
        return jsonify(error="Invalid student_type."), 400
    if student_type == 'day_scholar' and transport_type not in ('own_transport', 'college_bus'):
        return jsonify(error="Day scholar requires transport_type."), 400

    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT id, student_type, transport_type FROM users WHERE id=%s AND role='student'", (student_id,))
            stu = c.fetchone()
            if not stu: return jsonify(error="Student not found."), 404

            old_type      = stu['student_type']
            old_transport = stu['transport_type']
            type_changed  = (old_type != student_type or old_transport != transport_type)

            c.execute("""
                UPDATE users
                SET name=%s, branch=%s, year_level=%s,
                    student_type=%s, transport_type=%s, email=%s, parent_phone=%s
                WHERE id=%s AND role='student'
            """, (name, branch, year_level, student_type,
                  transport_type if student_type == 'day_scholar' else None,
                  email, parent_phone, student_id))

            if type_changed:
                # Close previous type history record
                c.execute("""
                    UPDATE student_type_history SET to_date=CURDATE()
                    WHERE student_id=%s AND to_date IS NULL
                """, (student_id,))
                # Insert new record
                c.execute("""
                    INSERT INTO student_type_history
                        (student_id, student_type, transport_type, from_date, changed_by)
                    VALUES (%s,%s,%s,CURDATE(),%s)
                """, (student_id, student_type,
                      transport_type if student_type == 'day_scholar' else None,
                      session.get('user')))

        conn.commit()
        audit('EDIT_STUDENT', 'student', student_id, {'name': name, 'type_changed': type_changed})
    finally: conn.close()
    return jsonify(success=True)


@app.route('/api/admin/bulk_add_students', methods=['POST'])
@write_required
def bulk_add_students():
    data             = request.json or {}
    college_code     = str(data.get('college_code') or '').strip()
    branch_code      = str(data.get('branch_code') or '').strip().lower()
    branch           = str(data.get('branch') or '').strip()
    year_level       = int(data.get('year_level', 0))
    student_type     = str(data.get('student_type') or '').strip()
    transport_type   = str(data.get('transport_type') or '').strip() or None
    roll_numbers_raw = str(data.get('roll_numbers') or '').strip()
    batch_id         = data.get('batch_id')

    if not college_code or not branch_code or not branch:
        return jsonify(error="college_code, branch_code and branch are required."), 400
    if year_level not in (1, 2, 3):
        return jsonify(error="year_level must be 1, 2, or 3."), 400
    if student_type not in ('hosteler', 'day_scholar'):
        return jsonify(error="Invalid student_type."), 400
    if student_type == 'day_scholar' and transport_type not in ('own_transport', 'college_bus'):
        return jsonify(error="Day scholar requires transport_type."), 400
    if not roll_numbers_raw:
        return jsonify(error="Roll numbers are required."), 400

    # Parse comma-separated numbers from frontend
    try:
        numbers = [int(x.strip()) for x in roll_numbers_raw.split(',') if x.strip()]
    except ValueError:
        return jsonify(error="Invalid roll numbers received."), 400

    if not numbers:
        return jsonify(error="No valid roll numbers provided."), 400
    if len(numbers) > 500:
        return jsonify(error="Max 500 students per bulk add."), 400

    # Build roll_no list
    roll_list = [f"{college_code}-{branch_code}-{n:03d}" for n in numbers]

    lock_err = _check_roll_lock(college_code)
    if lock_err: return lock_err

    created = skipped = 0
    results = []
    conn = get_db()
    try:
        with conn.cursor() as c:
            for roll_no in roll_list:
                c.execute("SELECT id FROM users WHERE roll_no=%s", (roll_no,))
                if c.fetchone():
                    skipped += 1
                    continue
                c.execute("""
                    INSERT INTO users
                        (roll_no, name, branch, username, year_level, password,
                         role, status, student_type, transport_type, batch_id, created_by)
                    VALUES (%s,%s,%s,%s,%s,%s,'student','active',%s,%s,%s,%s)
                """, (
                    roll_no, f"Student {roll_no}", branch, roll_no, year_level,
                    generate_password_hash(roll_no),
                    student_type,
                    transport_type if student_type == 'day_scholar' else None,
                    batch_id, session.get('user_id')
                ))
                uid = conn.insert_id()
                c.execute("""
                    INSERT INTO student_type_history
                        (student_id, student_type, transport_type, from_date, changed_by)
                    VALUES (%s,%s,%s,CURDATE(),%s)
                """, (uid, student_type,
                      transport_type if student_type == 'day_scholar' else None,
                      session.get('user')))
                results.append({'id': uid, 'roll_no': roll_no})
                created += 1

        conn.commit()
        if results:
            _auto_assign_fees(results, year_level, student_type, transport_type, batch_id)
        audit('BULK_ADD_STUDENTS', 'student', None,
              {'created': created, 'skipped': skipped})
    finally:
        conn.close()

    return jsonify(success=True, created=created, skipped=skipped)


def _check_roll_lock(roll_prefix):
    """Check if the roll number prefix is locked due to a graduated batch."""
    m = re.match(r'^(\d{2})', str(roll_prefix))
    if not m: return None
    batch_code = m.group(1)
    # Extract college_code prefix (first numeric segment)
    college_m = re.match(r'^(\d+)', str(roll_prefix))
    college_code = college_m.group(1) if college_m else roll_prefix

    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT id FROM graduated_batch_formats
                WHERE batch_code=%s
                AND (college_code=%s OR college_code LIKE %s)
            """, (batch_code, college_code, f"{batch_code}%"))
            if c.fetchone():
                return jsonify(
                    error=f"Roll number format starting with '{batch_code}' is permanently locked "
                          f"because a student from batch '{batch_code}' has graduated."
                ), 409
    finally: conn.close()
    return None


def _auto_assign_fees(students, year_level, student_type, transport_type, batch_id=None):
    """Auto-assign fees from active semesters applying fee rules:
    - Year 1 (Sem 1 only): ALL fee categories (college, building, hostel, bus) apply
    - Year 2 Sem 3 / Year 3 Sem 5: ALL fee categories apply
    - Year 2 Sem 4 / Year 3 Sem 6: ONLY hostel_fee and bus_fee apply
    Always filtered by student_type (hosteler/day_scholar).
    """
    sem_nos = DIPLOMA_SEMESTERS.get(year_level, [])
    if not sem_nos: return

    conn = get_db()
    try:
        with conn.cursor() as c:
            ph = ','.join(['%s'] * len(sem_nos))

            if batch_id:
                c.execute(f"""
                    SELECT s.id AS sem_id, s.semester_no FROM semesters s
                    JOIN academic_years ay ON s.academic_year_id=ay.id
                    WHERE s.semester_no IN ({ph}) AND ay.batch_id=%s
                    ORDER BY s.semester_no
                """, (*sem_nos, batch_id))
            else:
                c.execute(f"""
                    SELECT s.id AS sem_id, s.semester_no FROM semesters s
                    JOIN academic_years ay ON s.academic_year_id=ay.id
                    WHERE s.semester_no IN ({ph})
                    ORDER BY ay.is_active DESC, ay.id DESC
                    LIMIT {len(sem_nos)}
                """, tuple(sem_nos))

            sem_rows = c.fetchall()
            sem_ids  = [r['sem_id'] for r in sem_rows]
            sem_map  = {r['sem_id']: r['semester_no'] for r in sem_rows}

            if not sem_ids: return

            ph2 = ','.join(['%s'] * len(sem_ids))
            c.execute(f"""
                SELECT id, fee_name, amount, fee_category, semester_id
                FROM fee_structure WHERE semester_id IN ({ph2})
            """, tuple(sem_ids))
            fee_structures = c.fetchall()

            assigned = 0
            for stu in students:
                for fs in fee_structures:
                    cat     = (fs.get('fee_category') or '').lower()
                    sem_no  = sem_map.get(fs['semester_id'], 0)
                    is_even = sem_no in HOSTEL_BUS_ONLY_SEMS

                    # Even sems: skip college_fee and building_fee
                    if is_even and cat in ('college_fee', 'building_fee'):
                        continue
                    # Hostel fee → hostelers only
                    if cat == 'hostel_fee' and student_type != 'hosteler':
                        continue
                    # Bus fee → college_bus day scholars only
                    if cat == 'bus_fee' and student_type == 'hosteler':
                        continue
                    if cat == 'bus_fee' and student_type == 'day_scholar' \
                            and transport_type == 'own_transport':
                        continue

                    c.execute("""
                        SELECT id FROM student_fees
                        WHERE student_id=%s AND fee_structure_id=%s
                    """, (stu['id'], fs['id']))
                    if not c.fetchone():
                        c.execute("""
                            INSERT INTO student_fees (student_id, fee_structure_id, amount)
                            VALUES (%s,%s,%s)
                        """, (stu['id'], fs['id'], fs['amount']))
                        assigned += 1

        conn.commit()
        logger.info("Auto-assigned %d fees to %d students", assigned, len(students))
    finally: conn.close()

# ─────────────────────────────────────────────
#  API — PAYMENTS
# ─────────────────────────────────────────────
@app.route('/api/admin/add_payment', methods=['POST'])
@write_required
def add_payment():
    data             = request.json or {}
    err              = _require(data, 'student_id', 'fee_structure_id', 'amount',
                                'receipt_no', 'fee_type', 'date')
    if err: return err

    student_id       = int(data['student_id'])
    fee_structure_id = int(data['fee_structure_id'])
    amount           = float(data['amount'])
    receipt_no       = str(data['receipt_no']).strip()
    fee_type         = str(data['fee_type']).strip()
    pay_date         = str(data['date']).strip()
    payment_mode     = str(data.get('payment_mode') or 'cash').strip()
    installment_no   = data.get('installment_no')
    notes            = str(data.get('notes') or '').strip() or None
    admin_name       = session.get('user', 'Admin')

    if amount <= 0: return jsonify(error="Amount must be greater than zero."), 400
    if not receipt_no: return jsonify(error="Receipt number is required."), 400
    if payment_mode not in ('cash','dd','upi','online','cheque','scholarship'):
        return jsonify(error="Invalid payment mode."), 400

    try: datetime.strptime(pay_date, '%Y-%m-%d')
    except ValueError: return jsonify(error="Date must be YYYY-MM-DD."), 400

    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT id FROM payments WHERE receipt_no=%s", (receipt_no,))
            if c.fetchone():
                return jsonify(error=f"Receipt '{receipt_no}' already exists."), 409

            c.execute("""
                SELECT sf.amount, sf.waiver_amount, sf.scholarship_amount
                FROM student_fees sf
                WHERE sf.student_id=%s AND sf.fee_structure_id=%s
            """, (student_id, fee_structure_id))
            fee = c.fetchone()
            if not fee: return jsonify(error="Fee not assigned to this student."), 404

            net_due = float(fee['amount']) - float(fee['waiver_amount']) - float(fee['scholarship_amount'])

            c.execute("""
                SELECT IFNULL(SUM(amount),0) AS paid FROM payments
                WHERE student_id=%s AND fee_structure_id=%s AND status='Paid'
            """, (student_id, fee_structure_id))
            paid_so_far = float(c.fetchone()['paid'])

            if paid_so_far + amount > net_due + 0.01:
                due = net_due - paid_so_far
                return jsonify(error=f"Amount Rs.{amount} exceeds due Rs.{due:.2f}."), 400

            c.execute("""
                INSERT INTO payments
                    (student_id, fee_structure_id, amount, date,
                     status, receipt_no, fee_type, payment_mode,
                     recorded_by, installment_no, notes)
                VALUES (%s,%s,%s,%s,'Paid',%s,%s,%s,%s,%s,%s)
            """, (student_id, fee_structure_id, amount, pay_date,
                  receipt_no, fee_type, payment_mode, admin_name,
                  installment_no, notes))
            payment_id = conn.insert_id()

            # Update installment plan if applicable
            if installment_no:
                c.execute("""
                    UPDATE installment_plans SET paid_at=NOW(), payment_id=%s
                    WHERE student_id=%s AND fee_structure_id=%s AND installment_no=%s
                """, (payment_id, student_id, fee_structure_id, installment_no))

            c.execute("SELECT name, email, roll_no, parent_phone FROM users WHERE id=%s", (student_id,))
            stu = c.fetchone()

            # Get sem name for receipt
            c.execute("""
                SELECT s.semester_no, ay.year_name FROM fee_structure fs
                JOIN semesters s ON fs.semester_id=s.id
                JOIN academic_years ay ON s.academic_year_id=ay.id
                WHERE fs.id=%s
            """, (fee_structure_id,))
            sem_info = c.fetchone() or {}

        conn.commit()
        audit('RECORD_PAYMENT', 'payment', payment_id,
              {'receipt_no': receipt_no, 'amount': amount, 'student_id': student_id})

    finally: conn.close()

    # Generate PDF receipt
    pdf_bytes = None
    college = get_college_settings()
    if REPORTLAB_OK:
        pay_data = {
            'receipt_no': receipt_no, 'date': pay_date,
            'amount': amount, 'fee_type': fee_type,
            'payment_mode': payment_mode, 'recorded_by': admin_name,
            'sem_name': f"Sem {sem_info.get('semester_no','')}" if sem_info else '—',
        }
        pdf_bytes = generate_receipt_pdf(pay_data, stu or {}, college)

    # Email + WhatsApp
    if stu and stu.get('email'):
        send_receipt_email(stu['email'], stu['name'], stu['roll_no'],
                           receipt_no, fee_type, amount, pay_date,
                           admin_name, payment_mode, pdf_bytes)
    if stu and stu.get('parent_phone'):
        msg = (f"Fee Payment Confirmed\nReceipt: {receipt_no}\n"
               f"Student: {stu['name']} ({stu['roll_no']})\n"
               f"Amount: Rs.{amount:,.2f}\nDate: {pay_date}")
        send_whatsapp(stu['parent_phone'], msg)

    return jsonify(success=True, payment_id=payment_id)


@app.route('/api/admin/payments')
@admin_required
def get_payments():
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT
                    p.id, p.receipt_no, p.fee_type, p.recorded_by,
                    p.amount AS payment_amount, p.date, p.payment_mode,
                    p.installment_no, p.notes,
                    u.roll_no, u.name, u.branch, u.student_type,
                    fs.fee_name, s.semester_no, ay.year_name,
                    sf.amount AS total,
                    (sf.amount - sf.waiver_amount - sf.scholarship_amount) AS net_due,
                    IFNULL((
                        SELECT SUM(p2.amount) FROM payments p2
                        WHERE p2.student_id=p.student_id
                          AND p2.fee_structure_id=p.fee_structure_id
                          AND p2.status='Paid'
                    ),0) AS paid_total
                FROM payments p
                JOIN users u          ON p.student_id=u.id
                JOIN fee_structure fs  ON p.fee_structure_id=fs.id
                JOIN semesters s       ON fs.semester_id=s.id
                JOIN academic_years ay ON s.academic_year_id=ay.id
                JOIN student_fees sf
                    ON sf.student_id=p.student_id
                    AND sf.fee_structure_id=p.fee_structure_id
                WHERE p.status='Paid'
                ORDER BY p.id DESC
            """)
            rows = c.fetchall()
            for r in rows:
                r['payment_amount'] = float(r['payment_amount'])
                r['total']          = float(r['total'])
                r['net_due']        = float(r['net_due'])
                r['paid_total']     = float(r['paid_total'])
                r['due']            = round(r['net_due'] - r['paid_total'], 2)
                if r.get('date'): r['date'] = str(r['date'])
    finally: conn.close()
    return jsonify(rows)


@app.route('/api/admin/check_receipt/<path:receipt_no>')
@admin_required
def check_receipt(receipt_no):
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT id FROM payments WHERE receipt_no=%s", (receipt_no,))
            return jsonify(exists=bool(c.fetchone()))
    finally: conn.close()


@app.route('/api/admin/receipt_pdf/<int:payment_id>')
@admin_required
def download_receipt_pdf(payment_id):
    if not REPORTLAB_OK:
        return jsonify(error="PDF generation not available. Install reportlab."), 503
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT p.*, u.name, u.roll_no, u.branch, u.year_level, u.student_type,
                       fs.fee_name, s.semester_no, ay.year_name
                FROM payments p
                JOIN users u ON p.student_id=u.id
                JOIN fee_structure fs ON p.fee_structure_id=fs.id
                JOIN semesters s ON fs.semester_id=s.id
                JOIN academic_years ay ON s.academic_year_id=ay.id
                WHERE p.id=%s
            """, (payment_id,))
            p = c.fetchone()
    finally: conn.close()

    if not p: abort(404)
    college = get_college_settings()
    pay_data = {
        'receipt_no': p['receipt_no'], 'date': str(p['date']),
        'amount': float(p['amount']), 'fee_type': p['fee_type'],
        'payment_mode': p['payment_mode'], 'recorded_by': p['recorded_by'],
        'sem_name': f"Sem {p['semester_no']} — {p['year_name']}",
    }
    student_data = {
        'name': p['name'], 'roll_no': p['roll_no'],
        'branch': p['branch'], 'year_level': p['year_level'],
        'student_type': p['student_type'],
    }
    pdf = generate_receipt_pdf(pay_data, student_data, college)
    return send_file(io.BytesIO(pdf), mimetype='application/pdf',
                     as_attachment=True,
                     download_name=f"receipt_{p['receipt_no']}.pdf")


@app.route('/api/admin/no_dues_pdf/<int:student_id>')
@admin_required
def no_dues_pdf(student_id):
    if not REPORTLAB_OK:
        return jsonify(error="PDF generation not available. Install reportlab."), 503
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT * FROM users WHERE id=%s AND role='student'", (student_id,))
            stu = c.fetchone()
            if not stu: abort(404)
            c.execute("""
                SELECT SUM(sf.amount - sf.waiver_amount - sf.scholarship_amount -
                    IFNULL((SELECT SUM(p.amount) FROM payments p
                            WHERE p.student_id=sf.student_id
                            AND p.fee_structure_id=sf.fee_structure_id
                            AND p.status='Paid'),0)) AS total_due
                FROM student_fees sf WHERE sf.student_id=%s
            """, (student_id,))
            row = c.fetchone()
            total_due = float(row['total_due'] or 0)
    finally: conn.close()

    if total_due > 0.01:
        return jsonify(error=f"Student has pending dues of Rs.{total_due:,.2f}."), 400

    college = get_college_settings()
    pdf = generate_no_dues_pdf(stu, college)
    audit('NO_DUES_CERT', 'student', student_id)
    return send_file(io.BytesIO(pdf), mimetype='application/pdf',
                     as_attachment=True,
                     download_name=f"no_dues_{stu['roll_no']}.pdf")

# ─────────────────────────────────────────────
#  API — FEE STRUCTURE
# ─────────────────────────────────────────────
@app.route('/api/academic_years')
@admin_required
def get_academic_years():
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT ay.*, b.batch_year, b.college_code
                FROM academic_years ay
                LEFT JOIN batches b ON ay.batch_id=b.id
                ORDER BY ay.id DESC
            """)
            data = c.fetchall()
            for r in data:
                r['is_active'] = bool(r['is_active'])
    finally: conn.close()
    return jsonify(data)


@app.route('/api/academic_years/<int:year_id>/semesters')
@admin_required
def get_semesters(year_id):
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT s.*, ay.year_name
                FROM semesters s
                JOIN academic_years ay ON s.academic_year_id=ay.id
                WHERE s.academic_year_id=%s ORDER BY s.semester_no
            """, (year_id,))
            data = c.fetchall()
            for r in data:
                if r.get('due_date'): r['due_date'] = str(r['due_date'])
    finally: conn.close()
    return jsonify(data)


@app.route('/api/semesters/<int:sem_id>/set_due_date', methods=['POST'])
@write_required
def set_due_date(sem_id):
    data          = request.json or {}
    due_date      = data.get('due_date')
    late_fee_day  = float(data.get('late_fee_per_day', 0))
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                UPDATE semesters SET due_date=%s, late_fee_per_day=%s WHERE id=%s
            """, (due_date, late_fee_day, sem_id))
        conn.commit()
    finally: conn.close()
    return jsonify(success=True)


@app.route('/api/semesters/<int:sem_id>/fees')
@admin_required
def get_fee_structures(sem_id):
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT fs.*, s.semester_no, ay.year_name
                FROM fee_structure fs
                JOIN semesters s ON fs.semester_id=s.id
                JOIN academic_years ay ON s.academic_year_id=ay.id
                WHERE fs.semester_id=%s ORDER BY fs.id
            """, (sem_id,))
            data = c.fetchall()
            for r in data: r['amount'] = float(r['amount'])
    finally: conn.close()
    return jsonify(data)


@app.route('/api/semesters/<int:sem_id>/fees/bulk', methods=['POST'])
@write_required
def bulk_create_fees(sem_id):
    data  = request.json or {}
    fees  = data.get('fees', [])
    if not fees: return jsonify(error='No fees provided'), 400
    VALID = ('college_fee','building_fee','hostel_fee','bus_fee','exam_fee','library_fee','other')

    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT semester_no FROM semesters WHERE id=%s", (sem_id,))
            sem_row = c.fetchone()
            if not sem_row: return jsonify(error='Semester not found'), 404

            sem_no   = sem_row['semester_no']
            is_even  = sem_no in HOSTEL_BUS_ONLY_SEMS  # 4 or 6 only
            # Sem 1 is Year-1 only semester — all fee categories allowed
            # Sem 3,5 are first sems of Year 2&3 — all fee categories allowed
            # Sem 4,6 are second sems of Year 2&3 — hostel/bus only
            created  = 0

            for fee in fees:
                fee_name = str(fee.get('fee_name') or '').strip()
                amount   = float(fee.get('amount') or 0)
                category = str(fee.get('fee_category') or 'college_fee').strip()
                if not fee_name or amount <= 0: continue
                if category not in VALID: continue
                # Enforce even-sem rule
                if is_even and category in ('college_fee', 'building_fee'):
                    continue  # skip silently
                c.execute("""
                    INSERT INTO fee_structure (semester_id, fee_name, amount, fee_category)
                    VALUES (%s,%s,%s,%s)
                """, (sem_id, fee_name, amount, category))
                created += 1
        conn.commit()
    finally: conn.close()
    return jsonify(success=True, created=created)


@app.route('/api/fees/<int:fee_id>', methods=['DELETE'])
@super_required
def delete_fee_structure(fee_id):
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT COUNT(*) AS cnt FROM payments WHERE fee_structure_id=%s", (fee_id,))
            if c.fetchone()['cnt'] > 0:
                return jsonify(error="Cannot delete: payments already recorded for this fee."), 409
            c.execute("DELETE FROM student_fees WHERE fee_structure_id=%s", (fee_id,))
            c.execute("DELETE FROM fee_structure WHERE id=%s", (fee_id,))
        conn.commit()
    finally: conn.close()
    audit('DELETE_FEE', 'fee', fee_id)
    return jsonify(success=True)


@app.route('/api/semesters/<int:sem_id>/assign_fees', methods=['POST'])
@write_required
def assign_fees_to_students(sem_id):
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT * FROM fee_structure WHERE semester_id=%s", (sem_id,))
            fee_structures = c.fetchall()
            if not fee_structures: return jsonify(error='No fee structures for this semester'), 400

            c.execute("SELECT semester_no FROM semesters WHERE id=%s", (sem_id,))
            sem_row = c.fetchone()
            if not sem_row: return jsonify(error='Semester not found'), 404

            # Year 1 = Sem 1 only; Year 2 = Sem 3,4; Year 3 = Sem 5,6
            sem_to_year = {1:1, 3:2, 4:2, 5:3, 6:3}
            year_level  = sem_to_year.get(sem_row['semester_no'])
            is_even     = sem_row['semester_no'] in HOSTEL_BUS_ONLY_SEMS  # 4 or 6

            c.execute("""
                SELECT id, student_type, transport_type, year_level FROM users
                WHERE role='student' AND status='active' AND (%s IS NULL OR year_level=%s)
            """, (year_level, year_level))
            students = c.fetchall()

            assigned = skipped = 0
            for stu in students:
                for fs in fee_structures:
                    cat = (fs.get('fee_category') or '').lower()
                    if is_even and cat in ('college_fee','building_fee'):
                        continue
                    if cat == 'hostel_fee' and stu['student_type'] != 'hosteler':
                        continue
                    if cat == 'bus_fee' and stu['student_type'] == 'hosteler':
                        continue
                    if cat == 'bus_fee' and stu.get('transport_type') == 'own_transport':
                        continue
                    c.execute("""
                        SELECT id FROM student_fees
                        WHERE student_id=%s AND fee_structure_id=%s
                    """, (stu['id'], fs['id']))
                    if c.fetchone():
                        skipped += 1
                        continue
                    c.execute("""
                        INSERT INTO student_fees (student_id, fee_structure_id, amount)
                        VALUES (%s,%s,%s)
                    """, (stu['id'], fs['id'], fs['amount']))
                    assigned += 1
        conn.commit()
    finally: conn.close()
    return jsonify(success=True, assigned=assigned, skipped=skipped)

# ─────────────────────────────────────────────
#  API — FEE WAIVERS
# ─────────────────────────────────────────────
@app.route('/api/admin/add_waiver', methods=['POST'])
@write_required
def add_waiver():
    data             = request.json or {}
    err              = _require(data, 'student_id', 'fee_structure_id', 'waiver_type', 'waiver_amount')
    if err: return err

    student_id       = int(data['student_id'])
    fee_structure_id = int(data['fee_structure_id'])
    waiver_type      = str(data['waiver_type']).strip()
    waiver_amount    = float(data['waiver_amount'])
    reason           = str(data.get('reason') or '').strip()

    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT amount FROM student_fees
                WHERE student_id=%s AND fee_structure_id=%s
            """, (student_id, fee_structure_id))
            fee = c.fetchone()
            if not fee: return jsonify(error="Fee not assigned to student."), 404
            if waiver_amount > float(fee['amount']):
                return jsonify(error="Waiver cannot exceed total fee."), 400

            c.execute("""
                INSERT INTO fee_waivers
                    (student_id, fee_structure_id, waiver_type, reason, waiver_amount, granted_by)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, (student_id, fee_structure_id, waiver_type, reason,
                  waiver_amount, session.get('user')))
            c.execute("""
                UPDATE student_fees SET waiver_amount=waiver_amount+%s
                WHERE student_id=%s AND fee_structure_id=%s
            """, (waiver_amount, student_id, fee_structure_id))
        conn.commit()
        audit('ADD_WAIVER', 'student', student_id,
              {'type': waiver_type, 'amount': waiver_amount})
    finally: conn.close()
    return jsonify(success=True)

# ─────────────────────────────────────────────
#  API — SCHOLARSHIPS
# ─────────────────────────────────────────────
@app.route('/api/admin/add_scholarship', methods=['POST'])
@write_required
def add_scholarship():
    data             = request.json or {}
    err              = _require(data, 'student_id', 'fee_structure_id', 'scheme_name', 'amount')
    if err: return err

    student_id       = int(data['student_id'])
    fee_structure_id = int(data['fee_structure_id'])
    scheme_name      = str(data['scheme_name']).strip()
    amount           = float(data['amount'])

    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT s.id AS sem_id FROM fee_structure fs
                JOIN semesters s ON fs.semester_id=s.id
                WHERE fs.id=%s
            """, (fee_structure_id,))
            sem = c.fetchone()
            c.execute("""
                INSERT INTO scholarships
                    (student_id, fee_structure_id, scheme_name, amount, sem_id, verified_by)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, (student_id, fee_structure_id, scheme_name, amount,
                  sem['sem_id'] if sem else None, session.get('user')))
            c.execute("""
                UPDATE student_fees SET scholarship_amount=scholarship_amount+%s
                WHERE student_id=%s AND fee_structure_id=%s
            """, (amount, student_id, fee_structure_id))
        conn.commit()
        audit('ADD_SCHOLARSHIP', 'student', student_id,
              {'scheme': scheme_name, 'amount': amount})
    finally: conn.close()
    return jsonify(success=True)

# ─────────────────────────────────────────────
#  API — INSTALLMENTS
# ─────────────────────────────────────────────
@app.route('/api/admin/add_installment_plan', methods=['POST'])
@write_required
def add_installment_plan():
    data             = request.json or {}
    err              = _require(data, 'student_id', 'fee_structure_id', 'installments')
    if err: return err

    student_id       = int(data['student_id'])
    fee_structure_id = int(data['fee_structure_id'])
    installments     = data['installments']  # list of {amount, due_date}

    conn = get_db()
    try:
        with conn.cursor() as c:
            for i, inst in enumerate(installments, 1):
                c.execute("""
                    INSERT INTO installment_plans
                        (student_id, fee_structure_id, installment_no, amount, due_date)
                    VALUES (%s,%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE amount=%s, due_date=%s
                """, (student_id, fee_structure_id, i,
                      float(inst.get('amount', 0)), inst.get('due_date'),
                      float(inst.get('amount', 0)), inst.get('due_date')))
        conn.commit()
    finally: conn.close()
    return jsonify(success=True)


@app.route('/api/admin/installments/<int:student_id>/<int:fee_id>')
@admin_required
def get_installments(student_id, fee_id):
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT * FROM installment_plans
                WHERE student_id=%s AND fee_structure_id=%s
                ORDER BY installment_no
            """, (student_id, fee_id))
            rows = c.fetchall()
            for r in rows:
                r['amount'] = float(r['amount'])
                if r.get('due_date'): r['due_date'] = str(r['due_date'])
    finally: conn.close()
    return jsonify(rows)

# ─────────────────────────────────────────────
#  API — DEFAULTERS
# ─────────────────────────────────────────────
@app.route('/api/admin/defaulters')
@admin_required
def get_defaulters():
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT
                    u.id, u.roll_no, u.name, u.branch, u.year_level,
                    u.student_type, u.email, u.parent_phone,
                    s.semester_no, s.due_date, ay.year_name,
                    fs.fee_name, fs.fee_category,
                    (sf.amount - sf.waiver_amount - sf.scholarship_amount) AS net_due,
                    IFNULL((SELECT SUM(p.amount) FROM payments p
                            WHERE p.student_id=sf.student_id
                            AND p.fee_structure_id=sf.fee_structure_id
                            AND p.status='Paid'),0) AS paid,
                    (sf.amount - sf.waiver_amount - sf.scholarship_amount -
                     IFNULL((SELECT SUM(p.amount) FROM payments p
                             WHERE p.student_id=sf.student_id
                             AND p.fee_structure_id=sf.fee_structure_id
                             AND p.status='Paid'),0)) AS pending,
                    DATEDIFF(CURDATE(), s.due_date) AS days_overdue
                FROM student_fees sf
                JOIN users u ON sf.student_id=u.id
                JOIN fee_structure fs ON sf.fee_structure_id=fs.id
                JOIN semesters s ON fs.semester_id=s.id
                JOIN academic_years ay ON s.academic_year_id=ay.id
                WHERE u.role='student' AND u.status='active'
                AND s.due_date IS NOT NULL AND s.due_date < CURDATE()
                HAVING pending > 0
                ORDER BY days_overdue DESC, u.roll_no
            """)
            rows = c.fetchall()
            for r in rows:
                r['net_due'] = float(r['net_due'])
                r['paid']    = float(r['paid'])
                r['pending'] = float(r['pending'])
                if r.get('due_date'): r['due_date'] = str(r['due_date'])
    finally: conn.close()
    return jsonify(rows)


@app.route('/api/admin/send_reminders', methods=['POST'])
@write_required
def send_reminders():
    data        = request.json or {}
    student_ids = data.get('student_ids', [])
    conn = get_db()
    sent = failed = 0
    try:
        with conn.cursor() as c:
            for sid in student_ids:
                c.execute("""
                    SELECT u.name, u.email, u.roll_no, u.parent_phone,
                           s.semester_no, s.due_date, ay.year_name,
                           SUM(sf.amount - sf.waiver_amount - sf.scholarship_amount -
                               IFNULL((SELECT SUM(p.amount) FROM payments p
                                       WHERE p.student_id=sf.student_id
                                       AND p.fee_structure_id=sf.fee_structure_id
                                       AND p.status='Paid'),0)) AS pending
                    FROM users u
                    JOIN student_fees sf ON sf.student_id=u.id
                    JOIN fee_structure fs ON sf.fee_structure_id=fs.id
                    JOIN semesters s ON fs.semester_id=s.id
                    JOIN academic_years ay ON s.academic_year_id=ay.id
                    WHERE u.id=%s AND s.due_date < CURDATE()
                    GROUP BY u.id, s.id
                    HAVING pending > 0
                """, (sid,))
                rows = c.fetchall()
                for row in rows:
                    if row.get('email'):
                        ok = send_reminder_email(
                            row['email'], row['name'], row['roll_no'],
                            f"Sem {row['semester_no']} ({row['year_name']})",
                            str(row['due_date']), float(row['pending'])
                        )
                        if ok:
                            sent += 1
                            c.execute("""
                                INSERT INTO notification_log
                                    (student_id, type, purpose, status, sent_at)
                                VALUES (%s,'email','fee_reminder','sent',NOW())
                            """, (sid,))
                        else:
                            failed += 1
                    if row.get('parent_phone'):
                        msg = (f"Fee Reminder: {row['name']} ({row['roll_no']}) "
                               f"has pending fee of Rs.{float(row['pending']):,.2f} "
                               f"for Sem {row['semester_no']}. Due: {row['due_date']}")
                        send_whatsapp(row['parent_phone'], msg)
        conn.commit()
    finally: conn.close()
    audit('SEND_REMINDERS', 'batch', None, {'sent': sent, 'failed': failed})
    return jsonify(success=True, sent=sent, failed=failed)

# ─────────────────────────────────────────────
#  API — EXAM ELIGIBILITY
# ─────────────────────────────────────────────
@app.route('/api/admin/exam_eligibility/<int:sem_id>')
@admin_required
def check_exam_eligibility(sem_id):
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT
                    u.id, u.roll_no, u.name, u.branch, u.year_level,
                    SUM(sf.amount - sf.waiver_amount - sf.scholarship_amount -
                        IFNULL((SELECT SUM(p.amount) FROM payments p
                                WHERE p.student_id=sf.student_id
                                AND p.fee_structure_id=sf.fee_structure_id
                                AND p.status='Paid'),0)) AS pending
                FROM users u
                JOIN student_fees sf ON sf.student_id=u.id
                JOIN fee_structure fs ON sf.fee_structure_id=fs.id
                WHERE u.role='student' AND u.status='active' AND fs.semester_id=%s
                GROUP BY u.id
            """, (sem_id,))
            rows = c.fetchall()
            result = []
            for r in rows:
                pending = float(r['pending'] or 0)
                result.append({
                    'id':         r['id'],
                    'roll_no':    r['roll_no'],
                    'name':       r['name'],
                    'branch':     r['branch'],
                    'year_level': r['year_level'],
                    'pending':    pending,
                    'eligible':   pending <= 0,
                })
    finally: conn.close()
    return jsonify(result)

# ─────────────────────────────────────────────
#  API — TRANSFER CERTIFICATE
# ─────────────────────────────────────────────
@app.route('/api/admin/issue_tc', methods=['POST'])
@write_required
def issue_tc():
    data       = request.json or {}
    student_id = int(data.get('student_id') or 0)
    reason     = str(data.get('reason') or '').strip()

    if not student_id: return jsonify(error="student_id required."), 400
    conn = get_db()
    try:
        with conn.cursor() as c:
            # Check pending fees
            c.execute("""
                SELECT SUM(sf.amount - sf.waiver_amount - sf.scholarship_amount -
                    IFNULL((SELECT SUM(p.amount) FROM payments p
                            WHERE p.student_id=sf.student_id
                            AND p.fee_structure_id=sf.fee_structure_id
                            AND p.status='Paid'),0)) AS pending
                FROM student_fees sf WHERE sf.student_id=%s
            """, (student_id,))
            row = c.fetchone()
            pending = float(row['pending'] or 0)
            fees_cleared = pending <= 0

            if not fees_cleared:
                return jsonify(
                    error=f"Cannot issue TC. Student has pending dues of Rs.{pending:,.2f}.",
                    pending=pending
                ), 400

            c.execute("""
                INSERT INTO tc_records (student_id, issued_at, reason, issued_by, fees_cleared)
                VALUES (%s, CURDATE(), %s, %s, 1)
                ON DUPLICATE KEY UPDATE issued_at=CURDATE(), reason=%s, fees_cleared=1
            """, (student_id, reason, session.get('user'), reason))
            c.execute("""
                UPDATE users SET status='tc_issued', tc_issued=1, tc_date=CURDATE()
                WHERE id=%s
            """, (student_id,))
        conn.commit()
        audit('ISSUE_TC', 'student', student_id, {'reason': reason})
    finally: conn.close()
    return jsonify(success=True)

# ─────────────────────────────────────────────
#  API — PROMOTE STUDENTS
# ─────────────────────────────────────────────
@app.route('/api/promote_students/preview')
@admin_required
def promote_preview():
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT year_level, COUNT(*) AS count FROM users
                WHERE role='student' AND status='active'
                GROUP BY year_level ORDER BY year_level
            """)
            by_year = c.fetchall()
            c.execute("SELECT COUNT(*) AS v FROM users WHERE role='student' AND status='active'")
            total = c.fetchone()['v']
            c.execute("SELECT COUNT(*) AS v FROM users WHERE role='student' AND status='active' AND year_level=3")
            graduating = c.fetchone()['v']
            c.execute("SELECT ay.*, b.batch_year, b.college_code FROM academic_years ay LEFT JOIN batches b ON ay.batch_id=b.id ORDER BY ay.id DESC")
            years = c.fetchall()
            for y in years: y['is_active'] = bool(y['is_active'])
    finally: conn.close()
    return jsonify(total=total, graduating=graduating,
                   promoting=total - graduating, by_year=by_year, years=years)


@app.route('/api/promote_students', methods=['POST'])
@super_required
def promote_students():
    data        = request.json or {}
    new_year_id = data.get('new_year_id')
    if not new_year_id: return jsonify(error='new_year_id is required'), 400

    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT id FROM academic_years WHERE id=%s", (new_year_id,))
            if not c.fetchone(): return jsonify(error='Academic year not found'), 404

            # Get graduating students to lock their batch format
            c.execute("""
                SELECT u.roll_no, b.batch_code, b.college_code
                FROM users u
                LEFT JOIN batches b ON u.batch_id=b.id
                WHERE u.role='student' AND u.status='active' AND u.year_level=3
            """)
            graduating_students = c.fetchall()

            c.execute("""
                UPDATE users SET status='graduated'
                WHERE role='student' AND status='active' AND year_level=3
            """)
            graduated = conn.affected_rows()

            # Lock batch formats for graduated students
            for gs in graduating_students:
                if gs.get('batch_code') and gs.get('college_code'):
                    c.execute("""
                        INSERT IGNORE INTO graduated_batch_formats
                            (batch_code, college_code)
                        VALUES (%s,%s)
                    """, (gs['batch_code'], gs['college_code']))

            c.execute("""
                UPDATE users SET year_level=year_level+1
                WHERE role='student' AND status='active' AND year_level IN (1,2)
            """)
            promoted = conn.affected_rows()

            c.execute("UPDATE academic_years SET is_active=0")
            c.execute("UPDATE academic_years SET is_active=1 WHERE id=%s", (new_year_id,))

            try:
                c.execute("""
                    INSERT INTO promotion_log (promoted, graduated, new_year_id, promoted_at)
                    VALUES (%s,%s,%s,NOW())
                """, (promoted, graduated, new_year_id))
            except: pass

        conn.commit()
        audit('PROMOTE_STUDENTS', 'batch', None,
              {'promoted': promoted, 'graduated': graduated})
    finally: conn.close()
    return jsonify(success=True, promoted=promoted, graduated=graduated)


@app.route('/api/undo_promotion', methods=['POST'])
@super_required
def undo_promotion():
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                UPDATE users SET year_level=year_level-1
                WHERE role='student' AND status='active' AND year_level IN (2,3)
            """)
            demoted = conn.affected_rows()
            c.execute("""
                UPDATE users SET status='active', year_level=3
                WHERE role='student' AND status='graduated'
            """)
            restored = conn.affected_rows()
        conn.commit()
        audit('UNDO_PROMOTION', 'batch', None,
              {'demoted': demoted, 'restored': restored})
    finally: conn.close()
    return jsonify(success=True, demoted=demoted, restored=restored)

# ─────────────────────────────────────────────
#  API — ADMIN MANAGEMENT (super_admin only)
# ─────────────────────────────────────────────
@app.route('/api/admin/admins')
@login_required(role='super_admin')
@login_required(min_role='admin')        
def get_admins():
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT id, username, role, email, name, created_at
                FROM users WHERE role IN ('super_admin','admin','accountant','viewer')
                ORDER BY id
            """)
            return jsonify(c.fetchall())
    finally: conn.close()


@app.route('/api/admin/admins', methods=['POST'])
@login_required(role='super_admin')
def create_admin():
    data     = request.json or {}
    err      = _require(data, 'username', 'password', 'role', 'email')
    if err: return err

    username = str(data['username']).strip()
    password = str(data['password'])
    role     = str(data['role']).strip()
    email    = str(data['email']).strip()
    name     = str(data.get('name') or username).strip()

    if role not in ('admin','accountant','viewer'):
        return jsonify(error="Role must be: admin, accountant, or viewer."), 400
    ok, reason = is_strong_password(password)
    if not ok: return jsonify(error=reason), 400

    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT id FROM users WHERE username=%s", (username,))
            if c.fetchone(): return jsonify(error="Username already exists."), 409
            c.execute("""
                INSERT INTO users (username, password, role, email, name, created_by)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, (username, generate_password_hash(password), role, email, name,
                  session.get('user_id')))
        conn.commit()
        audit('CREATE_ADMIN', 'user', conn.insert_id(),
              {'username': username, 'role': role})
    finally: conn.close()
    return jsonify(success=True)


@app.route('/api/admin/admins/<int:admin_id>', methods=['DELETE'])
@login_required(role='super_admin')
def delete_admin(admin_id):
    if admin_id == session.get('user_id'):
        return jsonify(error="Cannot delete your own account."), 400
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("DELETE FROM users WHERE id=%s AND role != 'super_admin'", (admin_id,))
        conn.commit()
        audit('DELETE_ADMIN', 'user', admin_id)
    finally: conn.close()
    return jsonify(success=True)

# ─────────────────────────────────────────────
#  API — AUDIT LOG
# ─────────────────────────────────────────────
@app.route('/api/admin/audit_log')
@login_required(role='super_admin')
def get_audit_log():
    page    = int(request.args.get('page', 1))
    per     = 50
    offset  = (page - 1) * per
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT al.*, u.username FROM audit_log al
                LEFT JOIN users u ON al.admin_id=u.id
                ORDER BY al.created_at DESC LIMIT %s OFFSET %s
            """, (per, offset))
            rows = c.fetchall()
            for r in rows:
                if r.get('created_at'): r['created_at'] = str(r['created_at'])
            c.execute("SELECT COUNT(*) AS v FROM audit_log")
            total = c.fetchone()['v']
    finally: conn.close()
    return jsonify(rows=rows, total=total, page=page, per=per)

# ─────────────────────────────────────────────
#  API — EXCEL REPORT
# ─────────────────────────────────────────────
@app.route('/api/admin/excel_report')
@admin_required
def excel_report():
    if not OPENPYXL_OK:
        return jsonify(error="Excel export not available. Install openpyxl."), 503

    report_type = request.args.get('type', 'payments')  # payments|defaulters|students
    conn = get_db()
    try:
        with conn.cursor() as c:
            if report_type == 'payments':
                c.execute("""
                    SELECT p.receipt_no, u.roll_no, u.name, u.branch,
                           p.fee_type, p.amount, p.payment_mode, p.date,
                           p.recorded_by, s.semester_no, ay.year_name
                    FROM payments p
                    JOIN users u ON p.student_id=u.id
                    JOIN fee_structure fs ON p.fee_structure_id=fs.id
                    JOIN semesters s ON fs.semester_id=s.id
                    JOIN academic_years ay ON s.academic_year_id=ay.id
                    WHERE p.status='Paid' ORDER BY p.date DESC
                """)
                headers = ['Receipt No','Roll No','Name','Branch','Fee Type',
                           'Amount','Payment Mode','Date','Recorded By','Semester','Academic Year']
            elif report_type == 'defaulters':
                c.execute("""
                    SELECT u.roll_no, u.name, u.branch, u.year_level, u.student_type,
                           s.semester_no, s.due_date,
                           (sf.amount - sf.waiver_amount - sf.scholarship_amount -
                            IFNULL((SELECT SUM(p.amount) FROM payments p
                                    WHERE p.student_id=sf.student_id
                                    AND p.fee_structure_id=sf.fee_structure_id
                                    AND p.status='Paid'),0)) AS pending,
                           DATEDIFF(CURDATE(), s.due_date) AS days_overdue
                    FROM student_fees sf
                    JOIN users u ON sf.student_id=u.id
                    JOIN fee_structure fs ON sf.fee_structure_id=fs.id
                    JOIN semesters s ON fs.semester_id=s.id
                    WHERE u.role='student' AND u.status='active'
                    AND s.due_date IS NOT NULL AND s.due_date < CURDATE()
                    HAVING pending > 0 ORDER BY days_overdue DESC
                """)
                headers = ['Roll No','Name','Branch','Year','Type','Semester',
                           'Due Date','Pending Amount','Days Overdue']
            else:
                c.execute("""
                    SELECT u.roll_no, u.name, u.branch, u.year_level, u.student_type,
                           u.transport_type, u.email, u.status, b.batch_year
                    FROM users u LEFT JOIN batches b ON u.batch_id=b.id
                    WHERE u.role='student' ORDER BY u.roll_no
                """)
                headers = ['Roll No','Name','Branch','Year','Type','Transport',
                           'Email','Status','Batch Year']
            rows = c.fetchall()
    finally: conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.title()

    college = get_college_settings()
    ws.merge_cells('A1:I1')
    ws['A1'] = college.get('college_name', 'FeeFlow Report')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:I2')
    ws['A2'] = f"{report_type.title()} Report — Generated {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')

    header_fill = PatternFill(start_color='1a3a6b', end_color='1a3a6b', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(rows, 5):
        for col_idx, val in enumerate(row.values(), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='f0f4ff', end_color='f0f4ff', fill_type='solid')

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{report_type}_report.xlsx')

# ─────────────────────────────────────────────
#  COLLEGE SETTINGS
# ─────────────────────────────────────────────
@app.route('/api/admin/college_settings', methods=['GET'])
@admin_required
def api_get_college_settings():
    return jsonify(get_college_settings())


@app.route('/api/admin/college_settings', methods=['POST'])
@login_required(role='super_admin')
def api_update_college_settings():
    data = request.json or {}
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                UPDATE college_settings SET
                    college_name=%s, address=%s, phone=%s, email=%s,
                    principal_name=%s, website=%s
                WHERE id=1
            """, (
                data.get('college_name'), data.get('address'),
                data.get('phone'), data.get('email'),
                data.get('principal_name'), data.get('website')
            ))
        conn.commit()
        audit('UPDATE_COLLEGE_SETTINGS', 'settings', 1)
    finally: conn.close()
    return jsonify(success=True)

# ─────────────────────────────────────────────
#  STUDENT PORTAL
# ─────────────────────────────────────────────
@app.route('/student')
@login_required(role='student')
def student_portal():
    return render_template('student_portal.html',
                           student_name=session.get('user', 'Student'))


@app.route('/api/student/my_fees')
@login_required(role='student')
def my_fees():
    student_id = session.get('user_id')
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT u.*, b.batch_year FROM users u
                LEFT JOIN batches b ON u.batch_id=b.id
                WHERE u.id=%s
            """, (student_id,))
            student = c.fetchone()
            c.execute("""
                SELECT sf.*, fs.fee_name, fs.fee_category, s.semester_no,
                       ay.year_name, s.due_date,
                       IFNULL((SELECT SUM(p.amount) FROM payments p
                               WHERE p.student_id=sf.student_id
                               AND p.fee_structure_id=sf.fee_structure_id
                               AND p.status='Paid'),0) AS paid
                FROM student_fees sf
                JOIN fee_structure fs ON sf.fee_structure_id=fs.id
                JOIN semesters s ON fs.semester_id=s.id
                JOIN academic_years ay ON s.academic_year_id=ay.id
                WHERE sf.student_id=%s ORDER BY s.semester_no
            """, (student_id,))
            fees = c.fetchall()
            for f in fees:
                f['amount'] = float(f['amount'])
                f['paid']   = float(f['paid'])
                f['due']    = round(float(f['amount']) - float(f['waiver_amount'])
                                    - float(f['scholarship_amount']) - f['paid'], 2)
                if f.get('due_date'): f['due_date'] = str(f['due_date'])

            c.execute("""
                SELECT p.*, fs.fee_name, s.semester_no, ay.year_name
                FROM payments p
                JOIN fee_structure fs ON p.fee_structure_id=fs.id
                JOIN semesters s ON fs.semester_id=s.id
                JOIN academic_years ay ON s.academic_year_id=ay.id
                WHERE p.student_id=%s AND p.status='Paid'
                ORDER BY p.date DESC
            """, (student_id,))
            payments = c.fetchall()
            for p in payments:
                p['amount'] = float(p['amount'])
                if p.get('date'): p['date'] = str(p['date'])
    finally: conn.close()
    return jsonify(student=student, fees=fees, payments=payments)


@app.route('/api/student/download_receipt/<int:payment_id>')
@login_required(role='student')
def student_download_receipt(payment_id):
    student_id = session.get('user_id')
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT p.* FROM payments p
                WHERE p.id=%s AND p.student_id=%s
            """, (payment_id, student_id))
            if not c.fetchone(): abort(403)
    finally: conn.close()
    return download_receipt_pdf(payment_id)

# ─────────────────────────────────────────────
#  FORGOT PASSWORD / OTP / RESET
# ─────────────────────────────────────────────
@app.route('/forgot', methods=['GET', 'POST'])
@rate_limit(max_calls=5, window=60)
def forgot():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        if not email:
            flash("Please enter your email address.", "error")
            return render_template('forgot.html')
        conn = get_db()
        try:
            with conn.cursor() as c:
                c.execute("SELECT id FROM users WHERE email=%s AND role != 'student'", (email,))
                user = c.fetchone()
        finally: conn.close()
        session['pending_otp_email'] = email  # set BEFORE any redirect so verify_page works
        if user:
            otp = _generate_otp()
            _otp_save(email, otp)
            email_sent = send_otp_email(email, otp)
            if not email_sent:
                # OTP is stored in DB; warn but don't block — user can proceed if they received it
                logger.warning("OTP email delivery failed for %s", email)
                flash("OTP was generated but the email could not be delivered. "
                      "Please check your spam folder or try again.", "warning")
            else:
                flash("An OTP has been sent to your email address.", "info")
        else:
            # Don't reveal whether the email exists (security: enumeration prevention)
            flash("If that email is registered, an OTP has been sent.", "info")
        return redirect('/verify_page')
    return render_template('forgot.html')


@app.route('/verify_page')
def verify_page():
    if not session.get('pending_otp_email'):
        flash("Please start the password-reset process again.", "warning")
        return redirect('/forgot')
    return render_template('verify_otp.html', email=session['pending_otp_email'])


@app.route('/verify_otp', methods=['POST'])
@rate_limit(max_calls=10, window=60)
def verify_otp():
    email    = session.get('pending_otp_email', '').lower()
    user_otp = request.form.get('otp', '').strip()
    if not email:
        flash("Session expired. Please start again.", "error")
        return redirect('/forgot')
    record = _otp_get(email)
    if not record:
        flash("No OTP found. Please request a new one.", "error")
        return redirect('/forgot')
    age = (datetime.now() - record['created_at']).total_seconds()
    if age > OTP_EXPIRY_SECONDS:
        _otp_delete(email)
        flash("OTP expired. Please request a new one.", "error")
        return redirect('/forgot')
    if record['attempts'] >= OTP_MAX_ATTEMPTS:
        _otp_delete(email)
        flash("Too many failed attempts.", "error")
        return redirect('/forgot')
    if not secrets.compare_digest(str(record['otp']), user_otp):
        _otp_increment(email)
        remaining = max(0, OTP_MAX_ATTEMPTS - record['attempts'] - 1)
        flash(f"Incorrect OTP. {remaining} attempt(s) remaining.", "error")
        return redirect('/verify_page')
    _otp_delete(email)
    session['otp_verified_email'] = email
    session.pop('pending_otp_email', None)
    return redirect('/reset_password')


@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    email = session.get('otp_verified_email', '')
    if not email:
        flash("Unauthorized. Please complete OTP verification first.", "error")
        return redirect('/forgot')
    if request.method == 'POST':
        new_pw  = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')
        if new_pw != confirm:
            flash("Passwords do not match.", "error")
            return render_template('reset_password.html')
        ok, reason = is_strong_password(new_pw)
        if not ok:
            flash(reason, "error")
            return render_template('reset_password.html')
        conn = get_db()
        try:
            with conn.cursor() as c:
                c.execute("""
                    UPDATE users SET password=%s WHERE email=%s AND role != 'student'
                """, (generate_password_hash(new_pw), email))
                updated = conn.affected_rows()
            conn.commit()
        finally: conn.close()
        session.pop('otp_verified_email', None)
        if updated:
            flash("Password reset successfully. Please log in.", "success")
        else:
            flash("Could not update password.", "error")
        return redirect('/')
    return render_template('reset_password.html')
# Add near the bottom of app.py, before the error handlers

@app.route('/api/admin/record_payment', methods=['POST'])
@write_required
def record_payment_alias():
    return add_payment()

@app.route('/api/admin/export_excel')
@admin_required
def export_excel_alias():
    return excel_report()

@app.route('/api/admin/waivers_list')
@admin_required
def waivers_list():
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("""
                SELECT fw.id, u.roll_no, u.name, fs.fee_name,
                       'waiver' AS record_type,
                       fw.waiver_type, fw.reason, fw.waiver_amount, fw.granted_by,
                       NULL AS scheme_name
                FROM fee_waivers fw
                JOIN users u ON fw.student_id=u.id
                JOIN fee_structure fs ON fw.fee_structure_id=fs.id
                UNION ALL
                SELECT sch.id, u.roll_no, u.name, fs.fee_name,
                       'scholarship' AS record_type,
                       NULL AS waiver_type, NULL AS reason, sch.amount AS waiver_amount,
                       sch.verified_by AS granted_by, sch.scheme_name
                FROM scholarships sch
                JOIN users u ON sch.student_id=u.id
                JOIN fee_structure fs ON sch.fee_structure_id=fs.id
                ORDER BY id DESC
            """)
            rows = c.fetchall()
            for r in rows:
                r['waiver_amount'] = float(r['waiver_amount'])
    finally:
        conn.close()
    return jsonify(rows)

@app.route('/api/admin/create_admin', methods=['POST'])
@login_required(role='super_admin')
def create_admin_alias():
    return create_admin()

@app.route('/api/admin/toggle_admin/<int:admin_id>', methods=['POST'])
@login_required(role='super_admin')
def toggle_admin(admin_id):
    data = request.json or {}
    new_status = data.get('status', 'active')
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("UPDATE users SET status=%s WHERE id=%s AND role != 'super_admin'",
                      (new_status, admin_id))
        conn.commit()
        audit('TOGGLE_ADMIN', 'user', admin_id, {'status': new_status})
    finally:
        conn.close()
    return jsonify(success=True)

@app.route('/api/admin/delete_admin/<int:admin_id>', methods=['DELETE'])
@login_required(role='super_admin')
def delete_admin_alias(admin_id):
    return delete_admin(admin_id)

@app.route('/api/promote_students_preview')
@admin_required
def promote_students_preview_alias():
    return promote_preview()

@app.route('/api/admin/change_password', methods=['POST'])
@admin_required
def change_password():
    data = request.json or {}
    new_password = str(data.get('new_password') or '').strip()
    if not new_password:
        return jsonify(error="New password is required."), 400
    ok, reason = is_strong_password(new_password)
    if not ok:
        return jsonify(error=reason), 400
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("UPDATE users SET password=%s WHERE id=%s",
                      (generate_password_hash(new_password), session.get('user_id')))
        conn.commit()
        audit('CHANGE_PASSWORD', 'user', session.get('user_id'))
    finally:
        conn.close()
    return jsonify(success=True)

@app.route('/api/admin/admins/<int:admin_id>', methods=['DELETE'])
@login_required(role='super_admin')  
def delete_admin_by_id(admin_id):
    return delete_admin(admin_id)

@app.route('/api/batches/<int:batch_id>', methods=['DELETE'])
@super_required
def delete_batch(batch_id):
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute("SELECT COUNT(*) AS cnt FROM users WHERE batch_id=%s AND role='student'",
                      (batch_id,))
            if c.fetchone()['cnt'] > 0:
                return jsonify(error="Cannot delete batch with enrolled students."), 409
            c.execute("DELETE FROM batches WHERE id=%s", (batch_id,))
        conn.commit()
        audit('DELETE_BATCH', 'batch', batch_id)
    finally:
        conn.close()
    return jsonify(success=True)
# ─────────────────────────────────────────────
#  ERROR HANDLERS
# ─────────────────────────────────────────────
@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', code=404, msg="Page not found."), 404

@app.errorhandler(500)
def server_error(e):
    logger.error("Internal server error: %s", e)
    return render_template('error.html', code=500, msg="Something went wrong."), 500

# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(debug=debug_mode, host='127.0.0.1', port=5000)